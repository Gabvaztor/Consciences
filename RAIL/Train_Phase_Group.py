"""
Seguir estructura del código

-1º. Leemos el fichero CSV que se corresponde con el archivo salida de Aquiles.
-2º. Obtenemos las columnas: "Tipo Evento", "Contenido" e "Imagen".
-3º. Eliminamos, de esas columnas, todas las filas cuyo "Contenido" o "Tipo Evento" sea vacío  o "Tipo Evento" no sea
     "Cursor" excepto para las que contengan en la columna "Contenido" un string o substring "{ctrl}k{ctrl}k". De esta
     forma nos quedamos solo con los "Tipo Evento" --> Cursor y con los "Tipo Evento" --> Keystrokes cuyo "Contenido"
     contenga un string o substring "{ctrl}k{ctrl}k", es decir, se haya detectado que el analista quiere que se aprenda
     a partir de ese momento hasta la siguiente vez que se encuentra "{ctrl}k{ctrl}k".
     Esto se deberá modificar cuando tengamos una versión en la que necesitemos recoger keystrokes diferentes.
-4º. Se crea, a partir de las columnas filtradas anteriormente, un DataFrame para poder trabajar con él.
-5º. Se recogen la lista de acciones que se han realizado entre una un "{ctrl}k{ctrl}k" y otro "{ctrl}k{ctrl}k".
-6º. Se recoge, a partir del nombre de la imagen, el grupo de la imagen para la fase de entrenamiento (aprendizaje).
     Hay que tener en cuenta que cada acción puede tener asociado una imagen diferente (esa información está en el log
     de aquiles). En ese caso, para esa imagen hay que buscarle al grupo que pertenece (esto se hace asociando los ids
     (nombre de la imagen) de los archivos image_match (que es el log enriquecido) y del log de aquiles).
     Se debe retornar algo así por cada acción:
     [Posición 0 --> 0 si "Tipo Evento" es Cursor y 1 si es Keystrokes,
     Posición 1 --> (A partir de "Contenido") 0 si es click izquierdo ,1 si es click derecho o "la cadena string" si
     es Keystrokes,
     Posición 2 --> Primera coordenada,
     Posición 3 --> Segunda coordenada,
     Posición 4 --> Grupo de la imagen de la fila del contenido en la que está. Si esa fila no tiene imagen asociada,
     se utilizará el grupo de la imagen anterior]
-7º. Se transforman todas las acciones(por acción) en un formato adecuado. El formato para
     las acciones debe quedar:
     [tipo_de_click, coordenada x, coordenada y, grupo]
     El tipo_de_click será 0 si es "{LEFT MOUSE}" y 1 si es "{RIGHT MOUSE}".
     Así, el array resultante podría quedar así: [0,450,354,4] siendo el último elemento el grupo.
     Esta función es útil para manejar versiones del código. Así, para las versiones del código en las que necesitemos
     un tipo de información diferente, podemos modificar esta función y recoger lo que necesitemos.
-8º. Creación de los datos de entrenamiento con los labels (etiquetas):
     Crear una lista de listas (A) que contenga las entradas y salidas del conjunto de entrenamiento.
     La entrada (a) tiene la forma del paso anterior. La salida(b) es la siguiente acción del log de aquiles con esta
     forma: [1,451,355], siendo el primer elemento el tipo de click, el segundo la coordenada x y el tercer elemento la
     coordenada y.
     Para la siguiente posición del array, (b) será la entrada y la salida de esa entrada será la siguiente acción
     del log de aquiles y así sucesivamente hasta que se encuentre el siguiente fin de aprendizaje
     (end_event_learning_capture).
     Un ejemplo de esto podría ser:
     [[[0,450,354,grupo], [1,451,355]], [[1,451,355,grupo],[0,349,567]], (...)] --> Conjunto de acciones divididas entre
     sus entradas y salidas(labels). Esto servirá para el entrenamiento y testeo de la red.
-9º. Una función que se llama "get_batch" que recibe por parámetro el array del anterior y un número entero y
     devuelva un número de elementos de ese array (batches) igual al número entero dado. Además, debe existir una
     variable que guarde el índice en el que está. Si tenemos un array de 1000 elementos y cogemos 50 ( la primera vez
     que se llama a la función), cada vez que se llame a esa función, debe dar los siguientes 50 elementos
     (tener en cuenta las ventanas temporales).

@(sara.moreno)
@(gabriel.vazquez)
"""
# pip3 install pandas
import pandas as pd
# Para instalar tensorflow (versión CPU): Una vez tengáis Python versión > 3 --> poner en línea de comandos:
# pip3 install --upgrade tensorflow
# Para versión GPU:
# pip3 install --upgrade tensorflow-gpu
import tensorflow as tf
import numpy as np
# Para leer excel
# Para instalar: pip3 install openpyxl
import openpyxl
# Para obtener la firma de las imágenes
from image_match.goldberg import ImageSignature
# Para recoger por parámetro
import argparse
# Just disables the warning, doesn't enable AVX/FMA
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
# Para calcular el tiempo
import time as time

# El argumento -c debe ser el último elemento ya que se recoge un path, no un fullpath.
ap = argparse.ArgumentParser()
ap.add_argument("-i", "--log_aquiles", required = False,
    help = "FullPath del archivo Aquiles")
ap.add_argument("-t", "--log_enriquecido", required = False,
    help = "FullPath del log enriquecido")
ap.add_argument("-p", "--load_previous_train", required = False,
    help = "S si quieres que se cargue el anterior modelo (para la predicción). Por defecto es False (No cargar)")
ap.add_argument("-c", "--path_images_test", required = False,
    help = "Path de la carpeta que contiene la imagen para obtener su grupo a partir de la distancia hamming (para la predicción)")

args = vars(ap.parse_args())

aquiles_file = args["log_aquiles"]
enriched_file = args["log_enriquecido"]
load_previous_model = args["load_previous_train"]
images_aquiles_to_get_sign_path = args["path_images_test"]
if images_aquiles_to_get_sign_path:
    while images_aquiles_to_get_sign_path[-1] != "\"":
        images_aquiles_to_get_sign_path = images_aquiles_to_get_sign_path[:-1]
    images_aquiles_to_get_sign_path = images_aquiles_to_get_sign_path[:-1] + "\\"

if load_previous_model == "S":
    load_previous_model = True
else:
    load_previous_model = False

# VARIABLES GROBALES
# Keystrokes a tener en cuenta
start_event_learning_capture = "{ctrl}k{ctrl}k"
end_event_learning_capture = "{ctrl}k{ctrl}k"
# Path donde se guarda el modelo de la red neuronal
path_save_restore_model = "../RAIL/Model_Saved/model.ckpt"
# SE RECOGEN POR INPUT
# Path de Aquiles
if aquiles_file == None:
    aquiles_file = "C:\\Users\\Gabriel\\Desktop\\02. Aquiles\\dist Aquiles 20171113\\20180214_train_log_test_v1\\logfiles20180214.csv"
# Log enriquecido
if enriched_file == None:
    enriched_file = "C:\\Users\\Gabriel\\Desktop\\Proto1_v5\\image_match_first_test_la_escuela.xlsx"
"""
PARA LA FASE DE PREDICCIÓN
"""
# TODO Modificar en una futura versión
# Contiene la ruta de la carpeta donde está la imagen para obtener su grupo a partir de la distancia hamming
if images_aquiles_to_get_sign_path == None:
    images_aquiles_to_get_sign_path = "C:\\Users\Gabriel\Desktop\\02. Aquiles\dist Aquiles 20171113\\20180123\PRIMARY\\"


"""
Funciones útiles
"""
def numpy_fillna(data):
    """
    Esta función convierte todos los elementos en un array del mismo tamaño cada uno con ceros si faltan elementos.
    :param data:
    :return:
    """

    # Get lengths of each row of data
    lens = np.array([len(i) for i in data])

    # Mask of valid places in each row
    mask = np.arange(lens.max()) < lens[:,None]

    # Setup output array and put elements from data into masked positions
    out = np.zeros(mask.shape, dtype=data.dtype)
    out[mask] = np.concatenate(data)
    return out

def pt(title=None, text=None):
    """
    Use the print function to print a title and an object coverted to string
    :param title:
    :param text:
    """
    if text is None:
        text = title
        title = "-----------------------------"
    else:
        title += ':'
    print(str(title) + " \n " + str(text))

def initialize_session():
    """
    Initialize interactive session and all local and global variables
    :return: Session
    """
    sess = tf.InteractiveSession()
    sess.run(tf.local_variables_initializer())
    sess.run(tf.global_variables_initializer())
    return sess

def split_list_in_pairs (a_list):
    """
    A partir de una lista tal que "a_list" = [1,2,3,4,5,6,7], devuelve otra lista que contenga listas de elementos 2 a 2
    excepto si es impar que devolverá el último elemento en una lista propia.
    E.g.: de "a_list" devuelve --> [[1,2],[3,4],[5,6],[7]]
    """
    if len(a_list) == 1 or not a_list:
        final_list = [a_list]
    else:
        final_list = [a_list[i:i + 2] for i in range(0, len(a_list), 2)]
    return final_list
"""
PASOS
"""
# Paso 1º. Ruta del fichero
def read_file(name_file, separator):
    """
    Lee el fichero csv pasado como path
    """
    file = pd.read_csv(name_file, sep=separator)
    return file

# Paso 2º. Obtenemos las columnas: "Tipo Evento", "Contenido" e "Imagen"
def get_columns(file):
    """
    Se recogen las columnas "tipo_evento", "contenido" y "imagen" y se retornan
    """
    file = read_file(name_file=file, separator=";")
    tipo_evento = file.pop("Tipo Evento")
    contenido = file.pop("Contenido")
    imagen = file.pop(" Imagen")
    return tipo_evento,contenido,imagen

# Paso 3 y 4
def normalize_columns_and_create_dataframe(tipo_evento, contenido, imagen):
    """
    Paso 3. Eliminamos, de esas columnas, todas las filas cuyo "Contenido" o "Tipo Evento" sea vacío  o "Tipo Evento" no
    sea "Cursor" excepto para las que contengan en la columna "Contenido" un string o substring "{ctrl}k{ctrl}k". De
    esta forma nos quedamos solo con los "Tipo Evento" --> Cursor y con los "Tipo Evento" --> Keystrokes cuyo
    "Contenido" contenga un string o substring "{ctrl}k{ctrl}k", es decir, se haya detectado que el analista quiere que
    se aprenda a partir de ese momento hasta la siguiente vez que se encuentra "{ctrl}k{ctrl}k".
    Paso 4. Se crea, a partir de las columnas filtradas anteriormente, un DataFrame para poder trabajar con él y se
    retorna.
    """
    # Paso 3
    # Se comenta para no eliminar ningún elemento del log de Aquiles
    """
    for index in range(0, len(tipo_evento)):
        if (not tipo_evento[index] == "Cursor" or contenido[index] == "NaN" or contenido[index] == "") \
                and start_event_learning_capture not in contenido[index].lower():
            del tipo_evento[index]
            del contenido[index]
            del imagen[index]
    """
    # Paso 4
    dataset = {}
    dataset[0] = tipo_evento
    dataset[1] = contenido
    dataset[2] = imagen
    final_dataframe = pd.DataFrame(data=dataset)
    final_dataframe.columns = ['Tipo Evento', 'Contenido', 'Imagen']
    print(final_dataframe)
    return final_dataframe

# Paso 5. Obtenemos las imágenes a partir de los keystrokes y array de lista de acciones (índices)
def get_images_from_keystrokes(contenido):
    """
    Se recogen la lista de acciones que se han realizado entre una un "start_event_learning_capture" y otro
    "end_event_learning_capture".
    """
    #images_from_keystrokes = []  # Nombre de imágenes para utilizarlos como ids en el excel
    """
    # Obtenemos los índices de las filas con "Tipo Evento" --> Keystrokes
    keystrokes_indexes = [x for x in range(len(list(tipo_evento.values))) if list(tipo_evento.values)[x].lower() ==
                          "keystrokes"]
    """
    keystrokes_indexes = [x for x in range(len(list(contenido.values))) if  start_event_learning_capture in
                          list(contenido.values)[x].lower()]
    pt("keystrokes_indexes",keystrokes_indexes)
    starts_ends_indexes = split_list_in_pairs(
        a_list=keystrokes_indexes)  # Lista de parejas de todos los keystrokes_indexes
    print("keystrokes_indexes", keystrokes_indexes)
    print("index_in_pairs", starts_ends_indexes)
    #for index in keystrokes_indexes:
    #    images_from_keystrokes.append(list(imagen.values)[index])
    #print("images_from_keystrokes", images_from_keystrokes)
    return starts_ends_indexes


def get_click_and_coordinates(contenido_column):
    """
    Devuelve una lista que contiene:
    [contenido_information, coordinate_x, coordinate_y] a partir de la información de la fila y columna actual
    'contenido'.
    'contenido_information' podrá contener la información de tres formas:
    1- 0 si contenido contiene la cadena '{LEFT MOUSE}'
    2- 1 si contenido contiene la cadena '{RIGHT MOUSE}'
    3- El contenido en sí de la cadena ya que es keystroke. En este caso, coordinate_x y coordinate_y serán valdrán "-1"
    """
    contenido_information = contenido_column
    coordinate_x = -1000.
    coordinate_y = -1000.
    if "{LEFT MOUSE}" in contenido_column:
        contenido_information = -1 * 1000.
    elif "{RIGHT MOUSE}" in contenido_column:
        contenido_information = 1. * 1000.
    if contenido_information == (-1 * 1000.) or contenido_information == (1. * 1000.): # Si es un evento de ratón
        split_contenido_v1 = contenido_column.split(sep="}")
        split_contenido_v2 = split_contenido_v1[1].split(sep="-")
        coordinate_x = float(split_contenido_v2[0])
        coordinate_y = float(split_contenido_v2[1])
    return contenido_information, coordinate_x, coordinate_y

def get_group_from_image(image_id, path_enriched_log):
    """
    Obtiene, a partir de la cadena id de una imagen (el nombre), el grupo al que pertenece. Para ello, se busca su grupo
    en el log enriquecido.
    """
    group = -1  # Devuelve -1 si no encuentra el grupo.
    try:
        enriched_log = openpyxl.load_workbook(filename=path_enriched_log)
        sheet = enriched_log.get_sheet_by_name("Agrupacion")
        max_row = sheet.max_row
        row = None
        for row_index in range(1, max_row + 1):
            if sheet.cell(row=row_index, column=1).value == image_id:
                row = row_index
                break
        if row:
            group = float(sheet.cell(row=row, column=3).value) * 1000.
    except:
        ValueError("No se ha podido leer del log enriquecido")
    return group

# Paso 6. Obtenemos las siguientes acciones a realizar a partir de los índices obtenidos previamente.
# Se debe acceder al excel image_match con los ids de las imagenes para recoger el grupo de cada una de ellas.
def get_first_action(tipo_evento, contenido, imagen, path_enriched_log, current_pos):
    """
    Recogemos la acción anterior a la posición primera posición "start_event"
    "current_pos" contiene la posición del elemento "start_event"
    """
    action = None
    index_before = 1
    stop_flag = False
    image_id = None
    try:
        while start_event_learning_capture not in contenido[current_pos - index_before] and not stop_flag:
            actual_pos = current_pos - index_before
            event_type = None
            contenido_information = None
            coordinate_x = None
            coordinate_y = None
            if tipo_evento[actual_pos].lower() == "cursor":
                event_type = 0
            elif tipo_evento[actual_pos].lower() == "keystrokes":
                event_type = 1
            pt("event_type", event_type)
            contenido_information, coordinate_x, coordinate_y = get_click_and_coordinates(contenido[actual_pos])
            if not image_id:
                if str(type(imagen[actual_pos])) != "<class 'float'>":
                    if (imagen[actual_pos] != '' and
                                imagen[actual_pos].lower() != 'nan' and
                                imagen[actual_pos].lower() != None):
                        image_id = imagen[actual_pos]
            pt("contenido", contenido_information)
            pt("coordinate_x", coordinate_x)
            pt("coordinate_y", coordinate_y)
            pt("image_id and event_type and contenido_information and coordinate_x and coordinate_y",image_id and event_type and contenido_information and coordinate_x and coordinate_y)
            elements_list = [image_id, event_type, contenido_information, coordinate_x, coordinate_y]
            if [x for x in elements_list if x is not None] == elements_list:
                group_image = get_group_from_image(image_id, path_enriched_log)
                pt("group_image", group_image)
                if group_image != -1:
                    action = [event_type, contenido_information, coordinate_x, coordinate_y ,group_image]
                    pt("action", action)
                    stop_flag = True
            if not stop_flag:
                index_before += 1
    except:
        raise ValueError("No se ha conseguido obtener la acción anterior al start_event")
    return action
def get_next_actions(tipo_evento, contenido, imagen, path_enriched_log, starts_ends_indexes):
    next_actions = []  # Lista de listas de siguientes acciones. Cada lista contenida tiene el siguiente formato:
    # [Posición 0 --> 0 si "Tipo Evento" es Cursor y 1 si es Keystrokes,
    # Posición 1 --> (A partir de "Contenido") 0 si es click izquierdo ,1 si es click derecho o "la cadena string" si
    # es Keystrokes,
    # Posición 2 --> Primera coordenada,
    # Posición 3 --> Segunda coordenada,
    # Posición 4 --> Grupo de la imagen de la fila del contenido en la que está. Si esa fila no tiene imagen asociada,
    # se utilizará el grupo de la imagen anterior]
    # TODO Tener en cuenta la última acción antes del "start_event_learning_capture"
    if not starts_ends_indexes:
        next_actions = []  # No hay siguientes acciones
    else:
        for index in range(len(starts_ends_indexes)):
            first_action_flag = False
            current_learning = starts_ends_indexes[index]  # Contiene una lista con un número o un par de números
            pt("current_learning",current_learning)
            if (len(current_learning)) == 2:
                if (current_learning[1] -1) - (current_learning[0] -1) -1 >= 1:  # Si hay más de una acción (contando con la
                    # anterior)
                    # Añadimos una lista que contendrá las acciones para este intervalo entre entradas y salidas. Esta
                    # lista podrá ser vacía si no se cumple alguna de las lógicas de negocio de este método. Por ello,
                    # será necesario eliminar las listas que no contengan acciones.
                    next_actions.append([])
                    # Recogemos el id de la primera imagen disponible justo después de pulsar
                    # 'start_event_learning_capture'. Si no existe, imagen_id es vacío.
                    try:
                        image_id = imagen[current_learning[0]]
                    except:
                        image_id = None
                    for element_index in range(current_learning[0]+1,current_learning[1]):
                        """
                        pt("index", element_index+1)
                        pt("image_id", image_id)
                        pt("imagen[element_index]", imagen[element_index])
                        pt("type_imagen[element_index]",type(imagen[element_index]))
                        """
                        if not first_action_flag:
                            # Obtenemos la primera acción (justo la anterior a current_learning[0])
                            first_action = get_first_action(tipo_evento, contenido, imagen,
                                                            path_enriched_log=path_enriched_log,
                                                            current_pos=current_learning[0])
                            if first_action:
                                next_actions[index].append(first_action)
                                first_action_flag = True
                            else:
                                raise ValueError("No se ha podido obtener el primer elemento")
                        if first_action_flag:
                            if not image_id and not imagen[element_index]:
                                pt("Se actualiza image_id a None")
                                image_id = None
                            elif imagen[element_index] != image_id :
                                if str(type(imagen[element_index])) != "<class 'float'>":
                                    if (imagen[element_index] != '' and
                                         imagen[element_index].lower() != 'nan' and
                                         imagen[element_index].lower() != None):
                                            pt("imagen[element_index]",imagen[element_index])
                                            image_id = imagen[element_index]
                            if image_id:  # Si existe una imagen_id guardamos acción. Sino, pasamos a siguiente línea.
                                pt("image_id", image_id)
                                event_type = tipo_evento[element_index]
                                if event_type == "Cursor": # Si es cursor es 0, si es keystrokes es 1.
                                    event_type = 0
                                else:
                                    event_type = 1
                                # Obtenemos contenido_information (0 si es click izquierdo y 1 si es click derecho o el string
                                # en sí si es keystroke). Además las coordenadas si es un evento de cursor. Sino, las
                                # las coordenadas valen "-1".
                                contenido_information, coordinate_x, \
                                coordinate_y = get_click_and_coordinates(contenido[element_index])
                                # Obtiene el grupo de la imagen
                                group_image = get_group_from_image(image_id,path_enriched_log=path_enriched_log)
                                if group_image:  # Si existe grupo de la imagen (debe existir) TODO Analizar este if
                                    action = [event_type, contenido_information, coordinate_x, coordinate_y, group_image]
                                    pt("action",action)
                                    next_actions[index].append(action)
                        else: # No se ha conseguido obtener la acción previa al "start_event" por lo que no hay acción de
                            # partida
                            raise ValueError("No se ha conseguido obtener la acción previa al start_event" +
                                             "por lo que no hay acción de partida")
            else: # Solo tiene un elemento
                if len(next_actions) == 0:
                    next_actions = []  # No recogemos nada (en esta versión)
    return next_actions
# Paso 7
def normalize_actions(next_actions):
    """
    Elimina el primer elemento de cada lista de la lista de "next_actions".
    "next_actions" es de la forma:
    [actions_1,actions2,...] --> cada actions_x es de la forma:
    [action_1,action_2,...] --> cada action_x es de la forma:
    [Posición 0 --> 0 si "Tipo Evento" es Cursor y 1 si es Keystrokes,
     Posición 1 --> (A partir de "Contenido") 0 si es click izquierdo ,1 si es click derecho o "la cadena string" si
     es Keystrokes,
     Posición 2 --> Primera coordenada,
     Posición 3 --> Segunda coordenada,
     Posición 4 --> Grupo de la imagen de la fila del contenido en la que está. Si esa fila no tiene imagen asociada,
     se utilizará el grupo de la imagen anterior]
     En resumen, esta función eliminará la "Posición 0" de cada action_x.

     Además, para esta versión, eliminará los action_x que contengan keystrokes en vez de coordenadas.
    """
    # TODO Revisar versión
    recursive_flag = False
    for actions_index in range(len(next_actions)):
        length_actions = len(next_actions[actions_index])
        for action_index in range(length_actions):
            action_x = next_actions[actions_index][action_index]
            if not action_x:
                del next_actions[actions_index][action_index]
                recursive_flag = True
                break
            elif action_x[0] == 1:
                del next_actions[actions_index][action_index]
                recursive_flag = True
                break
            elif len(next_actions[actions_index][action_index]) == 5:
                del next_actions[actions_index][action_index][0]
    if recursive_flag:
        normalize_actions(next_actions=next_actions)
    #pt("next_actions_end",next_actions)
    return next_actions
# Paso 8
def create_batches(normalized_actions):
    """
    A partir de las acciones normalizadas, se crea el conjunto de entrenamiento. Este conjunto contendrá las entradas
    y salidas del las acciones normalizadas por cada iteración de acciones.

    Así, si existe por cada conjunto de acciones un total de 4 acciones, e.g.:
    [[0, 966.0, 591.0, 3], [0, 676.0, 267.0, 4], [0, 1359.0, 14.0, 5], [0, 731.0, 435.0, 6]]
    Se haría un conjunto de entrenamiento supervisado tal que:
    |--------------------ENTRADAS--------------------|--------------------SALIDAS--------------------|
                    [0, 966.0, 591.0, 3]----------------------------->[0, 676.0, 267.0]
                    [0, 676.0, 267.0, 4]----------------------------->[0, 1359.0, 14.0]
                    [0, 1359.0, 14.0, 5]----------------------------->[0, 731.0, 435.0]

    La dimensión del array sería, por cada conjunto de acciones y para este ejemplo: (3,3), es decir, 3 entradas y
    3 salidas.
    Las entradas tienen de dimensión (1,4), es decir, 1 fila y 4 columnas y las salidas tienen una dimensión (1,3),
    es decir, 1 fila y 3 columnas.
    En total, el tamaño total del array de entradas será de: (x,4), siendo x el número total de entradas y el array de
    salidas tendrá un tamaño de (x,3), siendo x el número total de salidas (que es el mismo que el de entradas).
    Se retornará un array en el que, accediendo a su primera posición, se retornarán las entradas y accediendo a su
    segunda posición se obtendrán las salidas. Hay que tener en cuenta que las entradas serán un conjunto de lista de
    acciones, por lo que accediendo a la primera posición de la primera posición del array se obtendrá la primera lista
    de acciones de entradas.

    Además, los números serán floats y
    """
    batches = []
    pt("normalized",normalized_actions[0])
    inputs = []
    outputs = []
    for actions_index in range(len(normalized_actions)):
        inputs.append([])
        outputs.append([])
        length_actions = len(normalized_actions[actions_index])
        for action_index in range(length_actions-1):
            action_input = normalized_actions[actions_index][action_index]
            action_output = normalized_actions[actions_index][action_index+1][:3]
            inputs[actions_index].append(np.asarray(action_input))
            outputs[actions_index].append(np.asarray(action_output))
    batches.append(np.asarray(inputs))
    batches.append(np.asarray(outputs))
    #pt("inputs", inputs)
    #pt("outputs", outputs)
    #pt("batches", batches)
    pt("batches.shape", len(batches))
    pt("batches[0]", np.asarray(batches[0]).shape)
    pt("batches[0][0]", np.asarray(batches[0][0]).shape)
    pt("batches[1]", np.asarray(batches[1]).shape)
    pt("batches[1][0]", np.asarray(batches[1][0]).shape)
    return batches

index_batch = 0  # Índice para tener en cuenta el índice por donde va el recorrido del batch.
index_total_batch = -1  # Índice que no puede sobrepasar "index_batch"
batches = []  # Que se utilizará para quedar guardada la primera vez que se cree.
# recoger como parámetro al igual que quantity y start_all_flag

# Paso 9
def get_batch(batches, quantity=None):
    """
    Devuelve x entradas y x salidas de la red donde x es 1 o "quantity".
    """
    global index_batch  # Para tener en cuenta el índice por donde va el recorrido del batch.
    # Si quantity es None, que recoja el total de acciones.
    global index_total_batch
    if quantity:
        # TODO Tener en cuenta quantity
        pass
    else:
        total_inputs_labels = np.asarray(batches[0]).shape[0] - 1
        index_total_batch = total_inputs_labels
        x_input = batches[0][index_batch]
        y_labels = batches[1][index_batch]
        if index_batch + 1 <= index_total_batch:
            index_batch += 1
        else:
            index_batch = 0
        return x_input, y_labels

def main_train_phase(log_aquiles_path=None, enriched_file=None, start_all_flag=False, quantity=None):
    """
    Este método realiza todos las fases para retornar un batch de tamaño "quantity".
    Si "start_all_flag" es False, realiza todos los pasos para crear el batch y guardarlo en memoria. Una vez que ese
    batch con todos los elementos esté creado (la primera vez o cuando se crea necesario realizarlo), se debe poner a
    False para solo retornar el batch de tamaño "quantity". Al realizar todos los pasos se actualiza la variable global
    "batches".
    Si en alguna de las fases se retorna un elemento vacío, el método mostrará por pantalla que no existen acciones a
    realizar y el elemento batches será un array vacío por lo que retornará un array vacío.
    """
    global batches
    if start_all_flag:
        tipo_evento, contenido, imagen = get_columns(file=log_aquiles_path)
        # normalize_columns_and_create_dataframe(tipo_evento=tipo_evento,contenido=contenido,imagen=imagen)
        starts_ends_indexes = get_images_from_keystrokes(contenido=contenido)
        pt("starts_ends_indexes", starts_ends_indexes)
        next_actions = get_next_actions(tipo_evento=tipo_evento,contenido=contenido, imagen=imagen,
                                        path_enriched_log=enriched_file,
                                        starts_ends_indexes=starts_ends_indexes)
        if not next_actions:
            pt("No hay acciones")
            raise ValueError("No hay acciones")
        else:
            normalized_actions = normalize_actions(next_actions=next_actions)
            batches = create_batches(normalized_actions)
    if not batches:
        pt("No hay elementos en el batch")
        raise ValueError("No hay elementos en el batch, trate de crear primero el batch para reutilizarlo")
    else:
        x_input, y_labels = get_batch(batches=batches,quantity=quantity)
    return x_input, y_labels

"""
RED NEURONAL
"""
def train_set_to_test():
    # Creación del conjunto de entrenamiento
    # input con label
    # Como se utiliza "mean_squared_error" para el cálculo del error, se estipula que 0 es click izquierdo y 1000 click
    # derecho. Así, el error absoluto buscará mejor de una forma más certera.
    tipo_click = 0.
    coordenada_x = 0.
    coordenada_y = 0.
    grupo = 0.

    input_batch = []
    label_batch = []

    for int in range(10):
        if int % 2 == 0:
            tipo_click = 1000.
            grupo = 1000.
            coordenada_x -= -50.
            coordenada_y -= -50.
        elif int % 3 == 0 :
            tipo_click = 1000.
            grupo = 2000.
        else:
            tipo_click = 0.
            grupo = 3000.
        coordenada_x+= 100.
        coordenada_y+= 100.

        input_batch.append([tipo_click,coordenada_x,coordenada_y,grupo])
        label_batch.append([tipo_click,coordenada_x,coordenada_y])

    inputs = np.asarray(input_batch).reshape(10,4)
    labels = np.asarray(label_batch)[::-1].reshape(10,3)

    test_set_input = np.asarray([0.,500.,500.,3000.])
    pt("inputs",inputs)
    pt("labels",labels)
    pt("test_set_input",test_set_input)
    return input_batch, label_batch, test_set_input

def create_restore_train_network(path_to_save_model, restore_flag=False, test_input=None):
    if restore_flag:
        tf.reset_default_graph()
    # Parametros de la red
    n_oculta_1 = 8 # 1ra capa de atributos
    n_oculta_2 = 8 # 2ra capa de atributos
    n_entradas = 4 # 4 datos de entrada
    n_clases = 3 # 3 salidas

    # input para los grafos
    x = tf.placeholder(tf.float32, [None, n_entradas],  name='DatosEntrada')
    y = tf.placeholder(tf.float32, [None, n_clases], name='Clases')

    # Creamos el modelo
    def perceptron_multicapa(x, pesos, sesgo):
        # Función de activación de la capa escondida
        capa_1 = tf.add(tf.matmul(x, pesos['h1']), sesgo['b1'])
        # activacion relu
        capa_1 = tf.nn.relu(capa_1)
        # Función de activación de la capa escondida
        capa_2 = tf.add(tf.matmul(capa_1, pesos['h2']), sesgo['b2'])
        # activación relu
        capa_2 = tf.nn.relu(capa_2)
        # Salida con activación lineal
        salida = tf.matmul(capa_2, pesos['out']) + sesgo['out']
        return salida


    # Definimos los pesos y sesgo de cada capa.
    pesos = {
        'h1': tf.Variable(tf.random_normal([n_entradas, n_oculta_1])),
        'h2': tf.Variable(tf.random_normal([n_oculta_1, n_oculta_2])),
        'out': tf.Variable(tf.random_normal([n_oculta_2, n_clases]))
    }
    sesgo = {
        'b1': tf.Variable(tf.random_normal([n_oculta_1])),
        'b2': tf.Variable(tf.random_normal([n_oculta_2])),
        'out': tf.Variable(tf.random_normal([n_clases]))
    }
    # Construimos el modelo
    pred = perceptron_multicapa(x, pesos, sesgo)

    # Definimos la funcion de coste
    error = tf.losses.mean_squared_error(labels=y,predictions=pred)
    # Algoritmo de optimización
    optimizar = tf.train.AdamOptimizer(learning_rate=0.01).minimize(error)

    sess = initialize_session()
    # Para guardar
    saver = tf.train.Saver()
    if not restore_flag:
        start_time = time.time()
        epochs = 15000
        trains = 10
        costs = []
        stop_train_flag = False
        min_error = 1000000000.
        save_path = None
        # Entrenamiento
        for epoca in range(epochs):
            if stop_train_flag:
                break
            for train in range(trains):
                # Get batch algorithm
                x_train, y_train = main_train_phase(aquiles_file, enriched_file, False)
                # Optimización por backprop y funcion de costo
                _, actual_error, y_ = sess.run([optimizar, error,pred],
                                         feed_dict={x: x_train, y: y_train})
                # Para debug
                if train == 0 and epoca % 10 == 0:
                    # imprimir información de entrenamiento
                    #pt("error",actual_error)
                    """
                    Para debug
                    """
                    #costs.append(actual_error)
                    #precisiones.append(accuracy)
                    #pt("x_train",x_train)
                    #pt("y_train",y_train)
                    #pt("y", y_)
                    pass
                if actual_error < 0.00001:  # Si error absoluto es menor a 0.1
                    pt("Actual error absoluto", actual_error)
                    min_error = actual_error
                    # Guarda las variables en un path
                    save_path = saver.save(sess, path_to_save_model)
                    stop_train_flag = True
                    break
            if epoca % 10 == 0:
                if actual_error < min_error:
                    pt('Tiempo', str(time.strftime("%Hh%Mm%Ss", time.gmtime((time.time() - start_time)))) + "\n Entrenando...")
                    pt("Actual error absoluto", actual_error)
                    min_error = actual_error
                    # Guarda las variables en un path
                    save_path = saver.save(sess, path_to_save_model)
                #pt("costes",costes)
                #pt("precisiones",precisiones)
        #pt("costes", costs)
        pt("FINAL para " + str(epochs) + " épocas con " + str(trains) + " entrenamientos cada una.")
        pt('Tiempo total de entrenamiento', str(time.strftime("%Hh%Mm%Ss", time.gmtime((time.time() - start_time)))))
        pt("Con un error absoluto de aprendizaje de ", min_error)
        pt("Modelo guardado en", save_path)
        return save_path
    else:
        # Restauramos el modelo
        saver.restore(sess, path_to_save_model)
        pt("Modelo restaurado con éxito")
        # Si existe conjunto de testeo, predecir valir
        if test_input is not None:
            prediction = pred.eval(feed_dict={x: test_input})
            pt("Predicción", prediction)

def get_path_from_robot_id_and_image_id(robot_id, image_id):
    """

    A partir del "robot_id" y de la "image_id" obtiene el path completo de la imagen.
    """
    # TODO
    return image_id


def get_image_sign_processed(path_image):
    """
    Obtiene la firma de la imagen dada y la procesa para que tenga el mismo formato que las firmas en el log
    enriquecido
    """
    image_signature = ImageSignature()
    sign_image = image_signature.generate_signature(path_image)
    sign_image_processed = str(sign_image).replace('\n', '*').replace('[ ', '').replace(']', '')
    # Procesamos la firma para que se obtenga de la misma forma que en el log enriquecido
    return sign_image_processed

def get_group_comparing_hamming_distance(image_sign, enriched_file_path):
    """

    A partir de la firma de una imagen, compara la firma con todas las firmas existentes en el log enriquecido y
    obtiene el grupo de la que haya obtenido la distancia hamming menor.
    """
    actual_group = None
    actual_hamming_distance = 900000000000
    def hamdist(str1, str2):
        """Count the # of differences between equal length strings str1 and str2"""
        diffs = 0
        for ch1, ch2 in zip(str1, str2):
            if ch1 != ch2:
                diffs += 1
        return diffs
    try:
        enriched_log = openpyxl.load_workbook(filename=enriched_file_path)
        sheet = enriched_log.get_sheet_by_name("Agrupacion")
        max_row = sheet.max_row
        for row_index in range(1, max_row + 1):
            if hamdist(sheet.cell(row=row_index, column=2).value, image_sign) < actual_hamming_distance:
                actual_hamming_distance = hamdist(sheet.cell(row=row_index, column=2).value, image_sign)
                actual_group = sheet.cell(row=row_index, column=3).value
        if actual_group:
            actual_group = float(actual_group) * 1000.
    except:
        ValueError("No se ha podido leer el log enriquecido")
    return actual_group
def get_group_from_image_id_and_hamming_distance(image_id, enriched_file_path, robot_id):
    """

    Obtiene el grupo de la imagen a partir de la "image_id", el "robot_id" y el log enriquecido

    """
    group = None
    # Path de la imagen para obtener la firma
    # TODO Para futura versión
    path_image = get_path_from_robot_id_and_image_id(robot_id, image_id)
    # TODO por ahora, recogemos el path real a partir de nuestro path de aquiles
    path_image = images_aquiles_to_get_sign_path + image_id
    # Firma de la imagen procesada
    image_sign = get_image_sign_processed(path_image=path_image)
    # Grupo de la imagen a través de la distancia hamming
    group = get_group_comparing_hamming_distance(image_sign=image_sign, enriched_file_path=enriched_file_path)
    if not group:
        raise ValueError("No se ha podido obtener el grupo de la imagen")
    return group


def get_last_action(tipo_evento, contenido, imagen, enriched_file_path, robot_id):
    """

    Recoge de las columnas "tipo_evento", "contenido" e "imagen" la última disponible acción disponible.
    A partir de la imagen, tenemos que buscar el grupo a partir del log enriquecido buscando la mínimia distancia
    hamming entre la última imagen y las firmas de las imágenes diponibles.

    """
    image_id = None
    contenido_data = None
    imagen = list(imagen.values)
    try:
        # Para coger la imagen
        for image_index in range(1, len(imagen)+2):
            if str(type(imagen[-image_index])) != "<class 'float'>":
                if (imagen[-image_index] != '' and
                            imagen[-image_index].lower() != 'nan' and
                            imagen[-image_index].lower() != None):
                    pt("imagen", imagen[-image_index])
                    image_id = imagen[-image_index]
                    break
        # TODO Comprobar versión. Para esta versión cogemos solo cuando es Cursor
        for contenido_index in range(1, len(tipo_evento.values)):
            if str(type(tipo_evento.values[-contenido_index])) != "<class 'float'>":
                if (tipo_evento.values[-contenido_index] != '' and
                            tipo_evento.values[-contenido_index].lower() != 'nan' and
                            tipo_evento.values[-contenido_index].lower() != None):
                    if tipo_evento.values[-contenido_index].lower() == "cursor":
                        contenido_data = contenido.values[-contenido_index]
                        pt("contenido_data", contenido_data)
                        break
        # Devuelve tres elementos:
        # "contenido_information" contiene -1000 si es click izquierdo, 1000 si es click derecho o una secuencia str.
        # "coordinate_x" y "coordinate_y" son las coordeandas. Valdrán -1000 si "contenido_information" es string.
        contenido_information, coordinate_x, coordinate_y = get_click_and_coordinates(contenido_data)
        # Obtenemos el grupo a partir de una "imagen_id"
        group = get_group_from_image_id_and_hamming_distance(image_id, enriched_file_path, robot_id)
        return [contenido_information, coordinate_x, coordinate_y, group]
    except:
        raise ValueError("No se ha podido obtener la última acción para la predicción")

def generate_test_input(log_aquiles_path, enriched_file_path, robot_id):
    """
    Recoge la última acción del log junto con la última imagen y se realiza la distancia hamming entre la firma de la
    imagen y las firmas de las iágenes del log enriquecido. El grupo de la acción será el grupo de la imagen cuya
    distancia hamming sea menor.
    Retorna esa última acción con el tipo de click, coordenadas y grupo de la imagen en un array.
    """

    tipo_evento, contenido, imagen = get_columns(file=log_aquiles_path)
    last_action = get_last_action(tipo_evento, contenido, imagen, enriched_file_path, robot_id)

    return np.asarray(last_action).reshape(1,4)

if __name__ == '__main__':
    restore_model = load_previous_model  # Para saltarse el entrenamiento y cargar el modelo (se debe tener el modelo guardado)
    pt("")
    pt("Comienzo del módulo 4")

    print("log_aquiles-->", aquiles_file)
    print("log_enriquecido-->", enriched_file)
    print("images_aquiles_to_get_sign_path-->", images_aquiles_to_get_sign_path)
    print("load_previous_model-->", load_previous_model)
    pt("")
    pt("¿Cargar modelo previo?", restore_model)
    # Get batch algorithm
    if not restore_model and not load_previous_model:
        main_train_phase(log_aquiles_path=aquiles_file, enriched_file=enriched_file, start_all_flag=True)
        save_path = create_restore_train_network(path_save_restore_model, restore_flag=False)
    else:
        test_set_input = generate_test_input(log_aquiles_path=aquiles_file, enriched_file_path=enriched_file,robot_id=2)
        pt("test_set_input", test_set_input)
        create_restore_train_network(path_save_restore_model, restore_flag=True, test_input=test_set_input)

